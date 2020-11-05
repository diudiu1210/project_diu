import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler:
    """Excel封装"""
    def __init__(self,file):
        # 初始化函数
        self.file=file

    def open_file(self,name) -> Worksheet:
        """打开表单"""
        wb=openpyxl.load_workbook(self.file)
        sheet=wb[name]
        return sheet

    def title(self,sheet_name):
        # 先调用开打开表单的方法
        sheet=self.open_file(sheet_name)
        # 读取标题就是读取第一行内容，把内容以一个列表存着，用for循环把第一行所有值打印出来添加到列表里
        # 首先定义一个空列表
        title=[]
        for column in sheet[1]:
            title.append(column.value)
        return title

    def total_test(self,sheet_name):
        """读取所有数据"""
        # 调用打开表单的方法
        sheet=self.open_file(sheet_name)
        # 获取所有数据固定用法变成列表，然后去掉标题行，从第二行开始，索引为1
        total_data=list(sheet.rows)[1:]
        # 定义一个空列表
        data=[]
        # 第一个for循环获取每一行
        for row in total_data:
            row_data = []
            for cell in row:
                # 把整行的数据值添加到列表里
                row_data.append(cell.value)
                # 用zip把标题和每一行的值一一对应起来，然后转化成字典
                data_dict=dict(zip(self.title(sheet_name),row_data))
            # 然后把字典添加到列表里
            data.append(data_dict)
        return data

    # Excel写入建议使用静态方法
    @staticmethod
    def write_cell(sheet_name,file,row,column,new_value):
        wb = openpyxl.load_workbook(file)
        sheet=wb[sheet_name]
        cell=sheet.cell(row,column)
        cell.value=new_value
        wb.save(file)
        wb.close()
if __name__ == '__main__':

    path=ExcelHandler(r"D:\diudiu\case.xlsx")
    # sheet_name=path.total_test("Sheet1")
    sheet_name=path.write_cell("Sheet1",r"D:\diudiu\case.xlsx",2,1,3333)
    print(sheet_name)

